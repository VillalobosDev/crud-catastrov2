import tkinter as tk
from tkinter import messagebox, filedialog
from tkinter import ttk
import customtkinter as ctk
from functions.functions import *
from modulos.menubar import menubar
from functions.calendario import open_calendar_popup
import customtkinter as ctk
from tkinter import ttk
from config.config import centrar_ventana
import openpyxl
import sqlite3
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

selected_id_liquidacion = None

def setup_treeview(frame):
    style = ttk.Style()
    style.configure("Custom.Treeview", font=("Poppins", 12), rowheight=25)
    style.configure("Custom.Treeview.Heading", font=("Poppins", 14, "bold"))

    treeview = ttk.Treeview(frame, style="Custom.Treeview", show="headings")
    treeview.pack(pady=10, padx=10, fill="both", expand=True)

    treeview["columns"] = ("CI", "Contribuyente", "Inmueble", "Monto 1", "Monto 2", "Fecha Liq 1", "Fecha Liq 2")
    for col in treeview["columns"]:
        treeview.heading(col, text=col.capitalize(), anchor="center")
        treeview.column(col, anchor="center")

    return treeview


def setup_treeview2(frame):
    style = ttk.Style()
    style.configure("Custom.Treeview", font=("Poppins", 12), rowheight=25)
    style.configure("Custom.Treeview.Heading", font=("Poppins", 14, "bold"))

    treeview = ttk.Treeview(frame, style="Custom.Treeview", show="headings")
    treeview.pack(pady=10, padx=10, fill="both", expand=True)

    treeview["columns"] = ("CI", "Contribuyente")
    for col in treeview["columns"]:
        treeview.heading(col, text=col.capitalize(), anchor="center")
        treeview.column(col, anchor="center")

    return treeview


def liqui_info(data):
    
    poppins30bold = ("Poppins", 25, "bold")
    poppins20bold = ("Poppins", 20, "bold")
    poppins14bold = ("Poppins", 14, "bold")
    poppins16bold = ("Poppins", 16, "bold")

    popup = ctk.CTkToplevel()
    popup.title("Liquidación")
    popup.geometry("600x500")
    popup.grab_set()
    popup.resizable(False, False)
    
    centrar_ventana(popup, 600, 500)

    bton_atras= ctk.CTkButton(popup, text="Atrás", command=popup.destroy, font=poppins14bold)
    bton_atras.pack(padx=10, pady=10, anchor="e", side="bottom")

    frame = ctk.CTkFrame(popup, corner_radius=15)
    frame.pack(fill="both", expand=True, padx=5, pady=5)
    
    # Crear un contenedor para frame2 y frame3
    
    top_container = ctk.CTkFrame(frame, corner_radius=15)
    top_container.pack(fill="x", padx=5, pady=5, side="top")
    

    frame2 = ctk.CTkFrame(top_container, corner_radius=15)
    frame2.pack(side="left", fill="both", expand=True, padx=5, pady=5)

    frame3 = ctk.CTkFrame(top_container, corner_radius=15)
    frame3.pack(side="left", fill="both", expand=True, padx=5, pady=5)

    frame4 = ctk.CTkFrame(frame, corner_radius=15)
    frame4.pack(fill="both", expand=True, padx=5, pady=5)
    
    
    
    text_top = ctk.CTkLabel(frame2, text="Contribuyente", font=poppins20bold)
    text_top.pack(padx=10, pady=5, side="top")
    
    contribuyente_value = ctk.CTkLabel(frame2, text=f"• {data[1]}", font=poppins16bold)
    contribuyente_value.pack(padx=20, pady=10, anchor="w")

    ci_value = ctk.CTkLabel(frame2, text=f"• {data[0]}", font=poppins16bold)
    ci_value.pack(padx=20, pady=10, anchor="w")
    
    
    text_top2 = ctk.CTkLabel(frame3, text="Inmueble", font=poppins20bold)
    text_top2.pack(padx=10, pady=10, side="top")
    
    inmueble_value = ctk.CTkLabel(frame3, text=f"• {data[2]}", font=poppins16bold)
    inmueble_value.pack(padx=20, pady=10, anchor="w")
    
    
    text_top3= ctk.CTkLabel(frame4, text="Pagos del inmueble", font=poppins20bold)
    text_top3.pack(pady=10, side="top")
    
    #frames de los pagos

    left_frame = ctk.CTkFrame(frame4)
    left_frame.pack(side="left", fill="both", expand=True, padx=5, pady=5)

    right_frame = ctk.CTkFrame(frame4)
    right_frame.pack(side="right", fill="both", expand=True, padx=5, pady=5)

    frame4.columnconfigure(0, weight=1)
    frame4.columnconfigure(1, weight=1)
    
    # Agrupar elementos en left_frame
    text_monto1 = ctk.CTkLabel(left_frame, text="Monto Liquidado:", font=poppins16bold)
    text_monto1.pack(padx=10, pady=10, anchor="w")

    monto1_value = ctk.CTkLabel(left_frame, text=f"• Monto a pagar: {data[3]}", font=poppins14bold)
    monto1_value.pack(padx=10, pady=5, anchor="w")

    fecha_liq1_value = ctk.CTkLabel(left_frame, text=f"• Fecha de pago: {data[5]}", font=poppins14bold)
    fecha_liq1_value.pack(padx=10, pady=5, anchor="w")

    # Agrupar elementos en right_frame
    text_monto2 = ctk.CTkLabel(right_frame, text="Imp derecho ocupación:", font=poppins16bold)
    text_monto2.pack(padx=10, pady=10, anchor="w")

    monto2_value = ctk.CTkLabel(right_frame, text=f"• Monto a pagar: {data[4]}", font=poppins14bold)
    monto2_value.pack(padx=10, pady=5, anchor="w")

    fecha_liq2_value = ctk.CTkLabel(right_frame, text=f"• Fecha de pago: {data[6]}", font=poppins14bold)
    fecha_liq2_value.pack(padx=10, pady=5, anchor="w")
    

def reload_treeviewsearch(treeview, ci_contribuyente):
    ci_contribuyente = ci_contribuyente.get()
    if not ci_contribuyente:
        messagebox.showwarning("Advertencia", "Por favor ingrese una cedula para buscar.")
        load_liquidaciones_data(treeview)
        return

    try:
        with connection() as conn:
            cursor = conn.cursor()
            sql = ''' 
            SELECT  l.id_liquidacion, c.ci_contribuyente, c.nombres || ' ' || c.apellidos AS contribuyente_nombre, i.nom_inmueble, l.monto_1, l.monto_2, l.fecha_Liquidacion_1, l.fecha_Liquidacion_2
            FROM liquidaciones l
            JOIN inmuebles i ON l.id_inmueble = i.id_inmueble
            JOIN contribuyentes c ON l.id_contribuyente = c.id_contribuyente
            WHERE c.ci_contribuyente = ?
            '''
            cursor.execute(sql, (ci_contribuyente,))
            results = cursor.fetchall()

            # Clear existing rows
            for row in treeview.get_children():
                treeview.delete(row)

            if not results:
                messagebox.showerror("Error", "No se ha encontrado la cédula del contribuyente.")
                load_liquidaciones_data(treeview)
                return

            # Insert updated rows
            for row in results:
                treeview.insert("", "end", iid=row[0], values=row[1:])
    except Exception as e:
        print(f"Error refreshing Treeview: {e}")

def load_liquidaciones_data(treeview):
    try:
        with connection() as conn:
            cursor = conn.cursor()
            sql = """
            SELECT l.id_liquidacion,  c.v_e || "-" || c.ci_contribuyente AS cedula_completa, c.nombres || ' ' || c.apellidos AS contribuyente_nombre, i.nom_inmueble, l.monto_1, l.monto_2, l.fecha_Liquidacion_1, l.fecha_Liquidacion_2
            FROM liquidaciones l
            JOIN inmuebles i ON l.id_inmueble = i.id_inmueble
            JOIN contribuyentes c ON l.id_contribuyente = c.id_contribuyente
            """
            cursor.execute(sql)
            results = cursor.fetchall()

            # Clear existing rows
            for row in treeview.get_children():
                treeview.delete(row)

            # Insert updated rows
            for row in results:
                treeview.insert("", "end", iid=row[0], values=row[1:])

    except Exception as e:
        print(f"Error fetching data: {e}")

def load_liquidaciones_data2(treeview):
    try:
        with connection() as conn:
            cursor = conn.cursor()
            sql = """
            SELECT c.ci_contribuyente, c.nombres || ' ' || c.apellidos AS contribuyente_nombre
            FROM contribuyentes c
            """
            cursor.execute(sql)
            results = cursor.fetchall()

            # Clear existing rows
            for row in treeview.get_children():
                treeview.delete(row)

            # Insert updated rows
            for row in results:
                treeview.insert("", "end", values=row)

    except Exception as e:
        print(f"Error fetching data: {e}")

def update_contribuyente_info(ci_entry, nombre_entry, inmueble_menu):
    ci_contribuyente = ci_entry.get()
    try:
        with connection() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT id_contribuyente, nombres || ' ' || apellidos FROM contribuyentes WHERE ci_contribuyente = ?", (ci_contribuyente,))
            contribuyente = cursor.fetchone()
            if contribuyente:
                id_contribuyente, nombre_completo = contribuyente
                nombre_entry.delete(0, tk.END)
                nombre_entry.insert(0, nombre_completo)

                cursor.execute("SELECT nom_inmueble FROM inmuebles WHERE id_contribuyente = ?", (id_contribuyente,))
                inmuebles = cursor.fetchall()
                inmueble_menu.configure(values=[inmueble[0] for inmueble in inmuebles])
            else:
                nombre_entry.delete(0, tk.END)
                inmueble_menu.configure(values=[""])
    except Exception as e:
        print(f"Error updating contribuyente info: {e}")

   
def clearentrys(ci_entry, nombre_entry, inmueble_menu, monto1_entry, monto2_entry, fecha1_entry, fecha2_entry):
        ci_entry.delete(0, tk.END)
        ci_entry.configure(placeholder_text="Cedula Contribuyente")

        nombre_entry.delete(0, tk.END)
        nombre_entry.configure(placeholder_text="Nombre Contribuyente")

        monto1_entry.delete(0, tk.END)
        monto1_entry.configure(placeholder_text="Monto 1")

        monto2_entry.delete(0, tk.END)
        monto2_entry.configure(placeholder_text="Monto 2")
        
        fecha1_entry.delete(0, tk.END)
        fecha1_entry.configure(placeholder_text="Fecha Liquidación 1")

        fecha2_entry.delete(0, tk.END)
        fecha2_entry.configure(placeholder_text="Fecha Liquidación 2")

        inmueble_menu.set("Inmuebles")

def update_liquidacion(tree, ci_entry, nombre_entry, inmueble_menu, monto1_entry, monto2_entry, fecha1_entry, fecha2_entry):
    global selected_id_liquidacion
    ci_contribuyente = ci_entry.get()
    inmueble = inmueble_menu.get()
    monto1 = monto1_entry.get()
    monto2 = monto2_entry.get()
    fecha1 = fecha1_entry.get()
    fecha2 = fecha2_entry.get()
    ci_contribuyente = ci_contribuyente[2:]
    print(ci_contribuyente)
    try:
        with connection() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT id_contribuyente FROM contribuyentes WHERE ci_contribuyente = ?", (ci_contribuyente,))
            contribuyente = cursor.fetchone()
            if contribuyente:
                id_contribuyente = contribuyente[0]
                cursor.execute("SELECT id_inmueble FROM inmuebles WHERE nom_inmueble = ? AND id_contribuyente = ?", (inmueble, id_contribuyente))
                inmueble = cursor.fetchone()
                if inmueble:
                    id_inmueble = inmueble[0]
                    cursor.execute("UPDATE liquidaciones SET monto_1 = ?, monto_2 = ?, fecha_Liquidacion_1 = ?, fecha_Liquidacion_2 = ?, id_inmueble = ? WHERE id_liquidacion = ?",
                                   (monto1, monto2, fecha1, fecha2, id_inmueble, selected_id_liquidacion))
                    conn.commit()
                    load_liquidaciones_data(tree)
                    clearentrys(ci_entry, nombre_entry, inmueble_menu, monto1_entry, monto2_entry, fecha1_entry, fecha2_entry)
                    messagebox.showinfo("Informacion", "Se ha actualizado la liquidación exitosamente")
                    print("Liquidación actualizada exitosamente.")
                else:
                    print("Inmueble no encontrado.")
            else:
                print("Error al actualizar el registro")
    except Exception as e:
        print(f"Error actualizando la liquidación: {e}")
        messagebox.showerror("Advertencia", f"Ha ocurrido un error: {e}")

def delete_liquidacion(ci_entry, inmueble_menu, my_tree):
    ci_contribuyente = ci_entry.get()
    inmueble = inmueble_menu.get()
    ci_contribuyente = ci_contribuyente[2:]

    confirm = messagebox.askyesno("Confirmación", "¿Está seguro de que desea eliminar esta liquidación?")
    if not confirm:
        return

    try:
        with connection() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT id_contribuyente FROM contribuyentes WHERE ci_contribuyente = ?", (ci_contribuyente,))
            contribuyente = cursor.fetchone()
            if contribuyente:
                id_contribuyente = contribuyente[0]
                cursor.execute("SELECT id_inmueble FROM inmuebles WHERE nom_inmueble = ? AND id_contribuyente = ?", (inmueble, id_contribuyente))
                inmueble = cursor.fetchone()
                if inmueble:
                    id_inmueble = inmueble[0]
                    cursor.execute("DELETE FROM liquidaciones WHERE id_inmueble = ?", (id_inmueble,))
                    conn.commit()
                    print("Liquidación eliminada exitosamente.")
                    messagebox.showinfo("Informacion", "Se a eliminado la liquidaición exitosamente")

                    load_liquidaciones_data(my_tree)
                else:
                    print("Inmueble no encontrado.")
            else:
                print("Error al eliminar la liquidacion")
    except Exception as e:
        print(f"Error eliminando la liquidación: {e}")

def liquidacion(window, last_window):
    global busquedabtn, busquedaliq, recargarbusqueda


    for widget in window.winfo_children():
        widget.destroy()

    poppins30bold = ("Poppins", 30, "bold")
    poppins20bold = ("Poppins", 20, "bold")
    poppins14bold = ("Poppins", 14, "bold")
    poppins12 = ("Poppins", 12, "bold")

    menubar(window)

    top_frame = ctk.CTkFrame(window, height=100, corner_radius=15)
    top_frame.pack(fill="x", padx=10, pady=10)

    top_frame2 = ctk.CTkFrame(window, height=100, corner_radius=15)
    top_frame2.pack(fill="x", padx=10)

    bottom_frame = ctk.CTkFrame(window, corner_radius=15)
    bottom_frame.pack(padx=10, pady=10, fill="both", expand=True)

    # Contenido del top frame
    menu = last_window
    volver_btn = ctk.CTkButton(top_frame, text="Volver", command=lambda: menu(window), font=poppins20bold)
    volver_btn.pack(padx=10, pady=10, side="left")

    window_title = ctk.CTkLabel(top_frame, text="Gestión de Liquidaciones", font=poppins30bold)
    window_title.pack(padx=10, pady=10, side="left")

    # Contenido del top frame 2
    
    recargarbusqueda = ctk.CTkButton(top_frame2, text="🔁", font=poppins14bold, width=30, command=lambda: load_liquidaciones_data(my_tree))
    recargarbusqueda.pack(padx=5, pady=5, side="right")


    busquedabtn = ctk.CTkButton(top_frame2, text="Buscar", font=poppins14bold, width=80, command=lambda: reload_treeviewsearch(my_tree, busquedaliq))
    busquedabtn.pack(padx=5, pady=5, side="right")

    busquedaliq = ctk.CTkEntry(top_frame2, placeholder_text="Buscar por cedula", font=poppins14bold, width=200)
    busquedaliq.pack(padx=5, pady=5, side="right")

    

    crearinm = ctk.CTkButton(top_frame2, text="Asignar", command=lambda: ifasignar(window, bottom_frame, top_frame2, busquedabtn, busquedaliq, last_window), font=poppins14bold)
    crearinm.pack(padx=5, pady=5, side="left")

    gestionarinm = ctk.CTkButton(top_frame2, text="Gestionar", command=lambda: ifgestionar(window, bottom_frame, top_frame2, busquedabtn, busquedaliq, last_window), font=poppins14bold)
    gestionarinm.pack(padx=5, pady=5, side="left")

    btn_exportar = ctk.CTkButton(top_frame2, text="Exportar Excel", command=exportar_a_excel, font=poppins14bold)
    btn_exportar.pack(padx=5, pady=5, side="left")



    def on_double_click(event):
        item = my_tree.selection()[0]
        data = my_tree.item(item, "values")
        liqui_info(data)
        

            
    # Contenido del bottom frame
    treeframe = ctk.CTkFrame(bottom_frame, corner_radius=15)
    treeframe.pack(padx=5, pady=5, fill="both", expand=True)

    # Creando el treeview para mostrar los registros
    frame_tree = ctk.CTkFrame(treeframe, fg_color='white', width=580, height=360)
    frame_tree.pack(pady=10, padx=10, expand=True, fill="both")

    my_tree = setup_treeview(frame_tree)
    load_liquidaciones_data(my_tree)

    # Configuración del estilo del Treeview (usando ttk dentro de CustomTkinter)
    style = ttk.Style()
    style.configure("Custom.Treeview", font=("Poppins", 12), rowheight=25)
    style.configure("Custom.Treeview.Heading", font=("Poppins", 14, "bold"))

    # Crear el scrollbar vertical con CustomTkinter
    horizontal_scrollbar = ttk.Scrollbar(frame_tree, orient="horizontal", command=my_tree.xview)
    my_tree.configure(xscrollcommand=horizontal_scrollbar.set)
    horizontal_scrollbar.pack(side="bottom", fill="x")

    # Añadir evento de doble clic
    my_tree.bind("<Double-1>", on_double_click)

def ifgestionar(window, bottom_frame, top_frame2, busquedabtnold, busquedaliqold, last_window):
    global busquedabtn, busquedaliq, recargarbusqueda

    if busquedabtnold:
        busquedabtnold.pack_forget()
    if busquedaliqold:
        busquedaliqold.pack_forget()
    if recargarbusqueda:
        recargarbusqueda.pack_forget()

    poppins14bold = ("Poppins", 14, "bold")
    poppins18 = ("Poppins", 18, "bold")

    for widget in bottom_frame.winfo_children():
        widget.destroy()

    



    frame_left = ctk.CTkFrame(bottom_frame, corner_radius=15, width=400)
    frame_left.pack(padx=5, pady=5, side="left", fill="both")

    frame_right = ctk.CTkFrame(bottom_frame, corner_radius=15)
    frame_right.pack(padx=5, pady=5, side="right", fill="both", expand=True)


    text = ctk.CTkLabel(frame_left, text="Gestionar Liquidación", font=poppins18, width=250)
    text.pack(padx=10, pady=10)


    # # Add UI elements for the left frame
    ci_frame = ctk.CTkFrame(frame_left)
    # ci_frame.pack(padx=10, pady=5, fill="x")

    nombre_frame = ctk.CTkFrame(frame_left)
    # nombre_frame.pack(padx=10, pady=5, fill="x")

    contframe = ctk.CTkFrame(frame_left)
    contframe.pack(padx=10, pady=5, fill="x")

    labeltitle = ctk.CTkLabel(contframe, text='Información del contribuyente', font=poppins14bold)
    labeltitle.pack(pady=5)

    contframe2 = ctk.CTkFrame(contframe, corner_radius=10, width=240, height=40)
    contframe2.pack(padx=10, pady=10)
    contframe2.pack_propagate(False)

    labeltitle2 = ctk.CTkLabel(contframe2, text='', font=poppins14bold)
    labeltitle2.pack(pady=10)

    #############################################

    monto1_frame = ctk.CTkFrame(frame_left)
    monto1_frame.pack(padx=10, pady=5, fill="x")

    monto2_frame = ctk.CTkFrame(frame_left)
    monto2_frame.pack(padx=10, pady=5, fill="x")

    fecha1_frame = ctk.CTkFrame(frame_left)
    fecha1_frame.pack(padx=10, pady=5, fill="x")

    fecha2_frame = ctk.CTkFrame(frame_left)
    fecha2_frame.pack(padx=10, pady=5, fill="x")

    inmueble_frame = ctk.CTkFrame(frame_left)
    inmueble_frame.pack(padx=10, pady=5, fill="x")

    ######################### ENTRYS

    ci_entry = ctk.CTkEntry(ci_frame, placeholder_text="Cedula Contribuyente", font=poppins14bold, width=250)
    # ci_entry.pack(pady=5, padx=5, side="left", fill="x", expand=True)

    nombre_entry = ctk.CTkEntry(nombre_frame, placeholder_text="Nombre Contribuyente", font=poppins14bold, width=250)
    # nombre_entry.pack(pady=5, padx=5, side="left", fill="x", expand=True)

    monto1 = ctk.CTkEntry(monto1_frame, placeholder_text="Monto 1", font=poppins14bold, width=250)
    monto1.pack(pady=5, padx=5, side="left", fill="x", expand=True)

    monto2 = ctk.CTkEntry(monto2_frame, placeholder_text="Monto 2", font=poppins14bold, width=250)
    monto2.pack(pady=5, padx=5, side="left", fill="x", expand=True)

    fecha1 = ctk.CTkEntry(fecha1_frame, placeholder_text="Fecha Liquidacion 1", font=poppins14bold, width=190)
    fecha1.pack(pady=5, padx=5, side="left", fill="x", expand=True)

    fecha1_btn = ctk.CTkButton(fecha1_frame, text="📅", command=lambda: open_calendar_popup(fecha1), font=poppins14bold, width=50)
    fecha1_btn.pack(pady=5, padx=5, side="left")

    fecha2 = ctk.CTkEntry(fecha2_frame, placeholder_text="Fecha Liquidacion 2", font=poppins14bold, width=190)
    fecha2.pack(pady=5, padx=5, side="left", fill="x", expand=True)

    fecha2_btn = ctk.CTkButton(fecha2_frame, text="📅", command=lambda: open_calendar_popup(fecha2), font=poppins14bold, width=50)
    fecha2_btn.pack(pady=5, padx=5, side="left")

    inmuebles = []
    inmueble_menu = ctk.CTkOptionMenu(inmueble_frame, values=inmuebles, font=poppins14bold, width=250)
    inmueble_menu.pack(pady=5, padx=5, side="left", fill="x", expand=True)

    ci_entry.bind("<FocusOut>", lambda e: update_contribuyente_info(ci_entry, nombre_entry, inmueble_menu))


    btnvolver = ctk.CTkButton(frame_left, text="Atrás", command=lambda: liquidacion(window, last_window), font=poppins14bold)
    btnvolver.pack(padx=10, pady=10, anchor="e", side="bottom")

    btndelete = ctk.CTkButton(frame_left, text="Eliminar", command=lambda: delete_liquidacion(ci_entry, inmueble_menu, my_tree), font=poppins14bold)
    btndelete.pack(padx=10, pady=10, anchor="e", side="bottom")

    btnsave = ctk.CTkButton(frame_left, text="Guardar", command=lambda: update_liquidacion(my_tree, ci_entry, nombre_entry, inmueble_menu, monto1, monto2, fecha1, fecha2), font=poppins14bold)
    btnsave.pack(padx=10, pady=10, anchor="e", side="bottom")

    # Add Treeview for the right frame.
    frame_tree = ctk.CTkFrame(frame_right, fg_color="white")
    frame_tree.pack(pady=10, padx=10, expand=True, fill="both")

    my_tree = setup_treeview(frame_tree)
    load_liquidaciones_data(my_tree)

    horizontal_scrollbar = ttk.Scrollbar(frame_tree, orient="horizontal", command=my_tree.xview)

    my_tree.configure(xscrollcommand=horizontal_scrollbar.set)

    horizontal_scrollbar.pack(side="bottom", fill="x")

    
    recargarbusqueda = ctk.CTkButton(top_frame2, text="🔁", font=poppins14bold, width=30, command=lambda: load_liquidaciones_data(my_tree))
    recargarbusqueda.pack(padx=5, pady=5, side="right")      

    busquedabtn = ctk.CTkButton(top_frame2, text="Buscar", font=poppins14bold, width=80, command=lambda: reload_treeviewsearch(my_tree, busquedaliq))
    busquedabtn.pack(padx=5, pady=5, side="right")

    busquedaliq = ctk.CTkEntry(top_frame2, placeholder_text="Buscar por cedula", font=poppins14bold, width=200)
    busquedaliq.pack(padx=5, pady=5, side="right")
    
    my_tree.bind("<<TreeviewSelect>>", lambda e: gestionar_liquidacion(inmuebles))  
    
    def gestionar_liquidacion(inmuebles):
        global selected_id_liquidacion
        selected_item = my_tree.selection()
        if selected_item:
            selected_iid = selected_item[0]
            selected_id_liquidacion = selected_iid
            print(selected_iid)
            item = my_tree.item(selected_item)
            values = item['values']
            ci_entry.delete(0, tk.END)
            ci_entry.insert(0, values[0])
            nombre_entry.delete(0, tk.END)
            nombre_entry.insert(0, values[1])
            labeltitle2.configure(text=values[1])

            try:
                with connection() as conn:
                    cursor = conn.cursor()
                    sql1 = '''SELECT id_contribuyente FROM contribuyentes WHERE ci_contribuyente = ?'''
                    cedula = values[0][2:]
                    cursor.execute(sql1, (cedula,))
                    result = cursor.fetchone()
                    if result:
                        id_contribuyente = result[0]
                        sql = "SELECT nom_inmueble FROM inmuebles WHERE id_contribuyente = ?"
                        cursor.execute(sql, (id_contribuyente,))
                        inmuebles_list = cursor.fetchall()
                        inmuebles.clear()
                        inmuebles.extend([inmueble[0] for inmueble in inmuebles_list])
                    else:
                        print("Contribuyente no encontrado.")
            except Exception as e:
                print(f'Error al cargar los inmuebles en el CtkOptionMenu: {e}')

            inmueble_menu.configure(values=inmuebles)
            if inmuebles:
                inmueble_menu.set(values[2])
            monto1.delete(0, tk.END)
            monto1.insert(0, values[3])
            monto2.delete(0, tk.END)
            monto2.insert(0, values[4])
            fecha1.delete(0, tk.END)
            fecha1.insert(0, values[5])
            fecha2.delete(0, tk.END)
            fecha2.insert(0, values[6])

            return selected_iid
    
def ifasignar(window, bottom_frame, top_frame2, busquedabtnold, busquedaliqold, last_window):
    global busquedabtn, busquedaliq, recargarbusqueda
    if busquedabtnold:
        busquedabtnold.pack_forget()
    if busquedaliqold:
        busquedaliqold.pack_forget()
    if recargarbusqueda:
        recargarbusqueda.pack_forget()

    poppins14bold = ("Poppins", 14, "bold")
    poppins18 = ("Poppins", 18, "bold")

    for widget in bottom_frame.winfo_children():
        widget.destroy()

    frame_left = ctk.CTkFrame(bottom_frame, corner_radius=15, width=400)
    frame_left.pack(padx=5, pady=5, side="left", fill="y")
    
    frame_right = ctk.CTkFrame(bottom_frame, corner_radius=15)
    frame_right.pack(padx=5, pady=5, side="right", fill="both", expand=True)

    text = ctk.CTkLabel(frame_left, text="Asignar Liquidación", font=poppins18, width=250)
    text.pack(padx=10, pady=10)

    ####################################################### Informacion del contribuyente

    contframe = ctk.CTkFrame(frame_left)
    contframe.pack(padx=10, pady=5, fill="x")

    labeltitle = ctk.CTkLabel(contframe, text='Información del contribuyente', font=poppins14bold)
    labeltitle.pack(pady=5)

    contframe2 = ctk.CTkFrame(contframe, corner_radius=10, width=240, height=40)
    contframe2.pack(padx=10, pady=10)
    contframe2.pack_propagate(False)

    labeltitle2 = ctk.CTkLabel(contframe2, text='', font=poppins14bold)
    labeltitle2.pack(pady=10)

    ############################################

    # Add UI elements for the left frame
    ci_frame = ctk.CTkFrame(frame_left)
    # ci_frame.pack(padx=10, pady=5, fill="x")

    nombre_frame = ctk.CTkFrame(frame_left)
    # nombre_frame.pack(padx=10, pady=5, fill="x")

    monto1_frame = ctk.CTkFrame(frame_left)
    monto1_frame.pack(padx=10, pady=5, fill="x")

    monto2_frame = ctk.CTkFrame(frame_left)
    monto2_frame.pack(padx=10, pady=5, fill="x")

    fecha1_frame = ctk.CTkFrame(frame_left)
    fecha1_frame.pack(padx=10, pady=5, fill="x")

    fecha2_frame = ctk.CTkFrame(frame_left)
    fecha2_frame.pack(padx=10, pady=5, fill="x")

    inmueble_frame = ctk.CTkFrame(frame_left)
    inmueble_frame.pack(padx=10, pady=5, fill="x")

    ##############################################

    ci_entry = ctk.CTkEntry(ci_frame, placeholder_text="Cedula Contribuyente", font=poppins14bold, width=250)
    # ci_entry.pack(pady=5, padx=5, side="left", fill="x", expand=True)

    nombre_entry = ctk.CTkEntry(nombre_frame, placeholder_text="Nombre Contribuyente", font=poppins14bold)
    # nombre_entry.pack(pady=5, padx=5, side="left", fill="x", expand=True)

    monto1 = ctk.CTkEntry(monto1_frame, placeholder_text="Monto 1", font=poppins14bold)
    monto1.pack(pady=5, padx=5, side="left", fill="x", expand=True)

    monto2 = ctk.CTkEntry(monto2_frame, placeholder_text="Monto 2", font=poppins14bold)
    monto2.pack(pady=5, padx=5, side="left", fill="x", expand=True)

    fecha1 = ctk.CTkEntry(fecha1_frame, placeholder_text="Fecha Liquidacion 1", font=poppins14bold)
    fecha1.pack(pady=5, padx=5, side="left", fill="x", expand=True)

    fecha1_btn = ctk.CTkButton(fecha1_frame, text="📅", command=lambda: open_calendar_popup(fecha1), font=poppins14bold, width=30)
    fecha1_btn.pack(pady=5, padx=5, side="left")

    fecha2 = ctk.CTkEntry(fecha2_frame, placeholder_text="Fecha Liquidacion 2", font=poppins14bold)
    fecha2.pack(pady=5, padx=5, side="left", fill="x", expand=True)

    fecha2_btn = ctk.CTkButton(fecha2_frame, text="📅", command=lambda: open_calendar_popup(fecha2), font=poppins14bold, width=30)
    fecha2_btn.pack(pady=5, padx=5, side="left")
    
    inmuebles = []
    inmueble_menu = ctk.CTkOptionMenu(inmueble_frame, values=inmuebles, font=poppins14bold, width=250)
    inmueble_menu.pack(pady=5, padx=5, side="left", fill="x", expand=True)

    ###################################### Add Treeview for the right frame
    frame_tree = ctk.CTkFrame(frame_right, fg_color="white")
    frame_tree.pack(pady=10, padx=10, expand=True, fill="both")
    
    ######################################
    my_tree = setup_treeview2(frame_tree)
    ######################################

    horizontal_scrollbar = ttk.Scrollbar(frame_tree, orient="horizontal", command=my_tree.xview)

    my_tree.configure(xscrollcommand=horizontal_scrollbar.set)

    horizontal_scrollbar.pack(side="bottom", fill="x")
    
    btnsave = ctk.CTkButton(frame_left, text="Guardar", command=lambda: save_liquidacion(my_tree, ci_entry, nombre_entry, inmueble_menu, monto1, monto2, fecha1, fecha2), font=poppins14bold)
    btnsave.pack(padx=10, pady=10, anchor="e", side="bottom")
    
    btnvolver = ctk.CTkButton(frame_left, text="Atrás", command=lambda: liquidacion(window, last_window), font=poppins14bold)
    btnvolver.pack(padx=10, pady=10, anchor="e", side="bottom")

    load_liquidaciones_data2(my_tree)
    
    recargarbusqueda = ctk.CTkButton(top_frame2, text="🔁", font=poppins14bold, width=30, command=lambda: load_liquidaciones_data2(my_tree))
    recargarbusqueda.pack(padx=5, pady=5, side="right")


    busquedabtn = ctk.CTkButton(top_frame2, text="Buscar", font=poppins14bold, width=80, command=lambda: reload_treeviewsearch(my_tree, busquedaliq))
    busquedabtn.pack(padx=5, pady=5, side="right")

    busquedaliq = ctk.CTkEntry(top_frame2, placeholder_text="Buscar por cedula", font=poppins14bold, width=200)
    busquedaliq.pack(padx=5, pady=5, side="right")

    my_tree.bind("<<TreeviewSelect>>", lambda e: on_tree_select(inmuebles))    


    def on_tree_select(inmuebles):
        global id_contr
        selected_item = my_tree.selection()
        if selected_item:
            item = my_tree.item(selected_item)
            values = item['values']
            labeltitle2.configure(text=f"{values[1]}")
            id_contr=f"{values[0]}"
            ci_entry.delete(0, tk.END)
            ci_entry.insert(0, values[0])
            nombre_entry.delete(0, tk.END)
            nombre_entry.insert(0, values[1])

        try:
            with connection() as conn:
                cursor = conn.cursor()
                sql1 = '''SELECT id_contribuyente FROM contribuyentes WHERE ci_contribuyente = ?'''
                cedula = values[0]
                cursor.execute(sql1, (cedula,))
                result = cursor.fetchone()
                if result:
                    id_contribuyente = result[0]
                    sql = "SELECT cod_catastral FROM inmuebles WHERE id_contribuyente = ?"
                    cursor.execute(sql, (id_contribuyente,))
                    inmuebles_list = cursor.fetchall()
                    inmuebles.clear()
                    inmuebles.extend([str(inmueble[0]) for inmueble in inmuebles_list])  # Convert to strings
                else:
                    print("Contribuyente no encontrado.")
        except Exception as e:
            print(f'Error al cargar los inmuebles en el CtkOptionMenu de ifasignar: {e}')

        inmueble_menu.configure(values=inmuebles)
        if inmuebles:
            inmueble_menu.set(inmuebles[0])

    def save_liquidacion(my_tree, ci_entry, nombre_entry, inmueble_menu, monto1_entry, monto2_entry, fecha1_entry, fecha2_entry):
        ci_contribuyente = ci_entry.get()
        inmueble = inmueble_menu.get()
        monto1 = monto1_entry.get()
        monto2 = monto2_entry.get()
        fecha1 = fecha1_entry.get()
        fecha2 = fecha2_entry.get()

        try:
            with connection() as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT id_contribuyente FROM contribuyentes WHERE ci_contribuyente = ?", (ci_contribuyente,))
                contribuyente = cursor.fetchone()
                if contribuyente:
                    id_contribuyente = contribuyente[0]
                    cursor.execute("SELECT id_inmueble FROM inmuebles WHERE cod_catastral = ? AND id_contribuyente = ?", (inmueble, id_contribuyente))
                    inmueble = cursor.fetchone()
                    if inmueble:
                        id_inmueble = inmueble[0]
                        cursor.execute("SELECT * FROM liquidaciones WHERE id_inmueble = ?", (id_inmueble,))
                        liquidacion = cursor.fetchone()
                        if liquidacion:
                            print("Ya existe una liquidación para este inmueble.")
                            messagebox.showwarning("Advertencia", "Ya existe una liquidación para este inmueble")
                        else:
                            cursor.execute("INSERT INTO liquidaciones (id_contribuyente, id_inmueble, monto_1, monto_2, fecha_Liquidacion_1, fecha_Liquidacion_2) VALUES (?, ?, ?, ?, ?, ?)",
                                           (id_contribuyente, id_inmueble, monto1, monto2, fecha1, fecha2))
                            conn.commit()
                            load_liquidaciones_data2(my_tree)
                            messagebox.showinfo("Informacion", "Se ha asignado la liquidación exitosamente")
                            print("Liquidación guardada exitosamente.")
                            my_tree.bind("<ButtonRelease-1>", on_tree_select)
                            clearentrys(ci_entry, nombre_entry, inmueble_menu, monto1_entry, monto2_entry, fecha1_entry, fecha2_entry)
                    else:
                        print("Inmueble no encontrado.")
                else:
                    print("Al guardar, Contribuyente no encontrado.")
        except Exception as e:
            print(f"Error guardando la liquidación: {e}")
 


def exportar_a_excel():
    try:
        # Conectar a la base de datos
        conn = sqlite3.connect('db.db')
        cursor = conn.cursor()

        # Consulta para obtener los datos
        cursor.execute('''
            SELECT l.fecha_Liquidacion_1, l.id_liquidacion, i.nom_inmueble, c.nombres || ' ' || c.apellidos, 
                   c.ci_contribuyente, s.nom_sector, i.cod_catastral, i.uso, l.monto_1, l.monto_2, l.fecha_Liquidacion_2
            FROM liquidaciones l
            JOIN inmuebles i ON l.id_inmueble = i.id_inmueble
            JOIN contribuyentes c ON l.id_contribuyente = c.id_contribuyente
            JOIN sectores s ON i.id_sector = s.id_sector
        ''')
        rows = cursor.fetchall()

        # Pedir al usuario que elija la ubicación y el nombre del archivo
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        if not file_path:
            return  # El usuario canceló la operación

        # Crear un nuevo archivo Excel
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Liquidaciones totales'

        # Agregar encabezados de columna
        headers = [
            "Fecha Liquidación", "N° Liquidación", "Inmueble", "Nombres y Apellidos",
            "Cédula", "Sector", "Código Catastral", "Uso", "Monto Liquidado",
            "Recargo del 10%", "Recargo del 15%", "Imp derecho de ocupación", "Total a Pagar", "Observación", "Fecha Liquidación"
        ]
        sheet.append(headers)

        # Cambiar la fuente y el tamaño de los encabezados
        header_font = Font(name='Arial', size=12, bold=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        
        for cell in sheet[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            
        # Resaltar las celdas de monto_2 y fecha_Liquidacion_2 en amarillo en los encabezados
        sheet["L1"].fill = yellow_fill  # monto_2
        sheet["O1"].fill = yellow_fill  # fecha_Liquidacion_2
            
            
        # Ajustar el ancho de las columnas
        column_widths = [30, 20, 25, 30, 20, 20, 30, 20, 30, 30, 30, 30, 30, 20, 30]
        for i, column_width in enumerate(column_widths, start=1):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = column_width

        # Agregar los datos al archivo Excel
        for row_data in rows:
            sheet.append([
                row_data[0],  # fecha_Liquidacion_1
                row_data[1],  # id_liquidacion
                row_data[2],  # nom_inmueble
                row_data[3],  # nombres y apellidos
                row_data[4],  # ci_contribuyente
                row_data[5],  # nom_sector
                row_data[6],  # cod_catastral
                row_data[7],  # uso
                row_data[8],  # monto_1
                "",           # Recargo del 10%
                "",           # Recargo del 15%
                row_data[9],  # monto_2
                row_data[8] + row_data[9],  # total_a_pagar
                "",           # Observación
                row_data[10]  # fecha_Liquidacion_2
            ])

        # Cambiar la fuente y el tamaño de los datos
        data_font = Font(name='Arial', size=10)
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.font = data_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
        # Resaltar las celdas de monto_2 y fecha_Liquidacion_2 en amarillo
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            row[11].fill = yellow_fill  # monto_2
            row[14].fill = yellow_fill  # fecha_Liquidacion_2
                
                
        # Resaltar las filas en rojo si las fechas están vacías
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            fecha_liquidacion_1 = row[0].value
            fecha_liquidacion_2 = row[14].value
            if not fecha_liquidacion_1 or not fecha_liquidacion_2:
                for cell in row:
                    cell.fill = red_fill
                    
                            
        # Agregar bordes a todas las celdas
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.border = thin_border

        # Guardar el nuevo archivo Excel en la ubicación seleccionada por el usuario
        workbook.save(file_path)
        messagebox.showinfo("Información", "Datos exportados exitosamente a Excel.")

    except PermissionError:
        messagebox.showerror("Error", "Permiso denegado: asegúrese de que el archivo no esté abierto en otra aplicación.")
    except Exception as e:
        messagebox.showerror("Error", f"Error al exportar los datos: {e}")

    finally:
        conn.close()





